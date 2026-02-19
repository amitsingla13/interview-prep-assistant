"""Generate the IT Helpdesk Knowledge Base Excel file."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "IT Helpdesk KB"

# Headers
headers = ["Incident_ID", "Category", "Sub_Category", "Issue_Description", "Resolution_Steps", "Priority", "Common_Cause", "Estimated_Time"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Knowledge Base Data - 80 realistic IT helpdesk incidents
incidents = [
    # === LAPTOP / DESKTOP ISSUES (1-15) ===
    ["INC001", "Laptop", "Won't Turn On", "Laptop does not power on when pressing the power button. No lights, no fan noise.",
     "1. Check if the charger is plugged in and the LED on the charger is lit.\n2. Try a different power outlet.\n3. Remove the battery (if removable), hold the power button for 30 seconds, reinsert battery.\n4. Try a hard reset: unplug charger, hold power button for 15 seconds.\n5. If still not working, try a different charger.\n6. If none of these work, raise a hardware ticket for motherboard/power issue.",
     "High", "Drained battery, faulty charger, or motherboard failure", "15-30 min"],

    ["INC002", "Laptop", "Slow Performance", "Laptop is extremely slow, applications take long to open, system freezes frequently.",
     "1. Restart the laptop (full shutdown, not just close lid).\n2. Open Task Manager (Ctrl+Shift+Esc) and check for high CPU/Memory/Disk usage.\n3. End any unnecessary processes consuming high resources.\n4. Clear temporary files: Win+R → type 'temp' → delete all files.\n5. Run Disk Cleanup: search 'Disk Cleanup' in Start menu.\n6. Check available disk space — ensure at least 10% free.\n7. Uninstall unused applications.\n8. Run a malware scan with Windows Defender.\n9. If issue persists, consider an SSD upgrade or RAM increase.",
     "Medium", "Too many startup programs, low disk space, malware, or insufficient RAM", "20-45 min"],

    ["INC003", "Laptop", "Blue Screen (BSOD)", "Laptop shows blue screen with error code and restarts automatically.",
     "1. Note the error code shown on the blue screen (e.g., IRQL_NOT_LESS_OR_EQUAL).\n2. Restart and check if issue recurs.\n3. Boot into Safe Mode: hold Shift while clicking Restart → Troubleshoot → Advanced → Startup Settings.\n4. In Safe Mode, uninstall recently installed software or drivers.\n5. Run 'sfc /scannow' in Command Prompt (admin) to check system files.\n6. Run 'DISM /Online /Cleanup-Image /RestoreHealth'.\n7. Check Windows Event Viewer for crash details.\n8. Update all drivers via Device Manager.\n9. If issue persists, back up data and consider OS reinstallation.",
     "High", "Driver conflicts, corrupted system files, hardware failure, or Windows updates", "30-60 min"],

    ["INC004", "Laptop", "Overheating", "Laptop becomes very hot, fans run constantly at high speed, may shut down unexpectedly.",
     "1. Place laptop on a hard, flat surface (not on bed/pillow).\n2. Check if air vents are blocked — clean with compressed air.\n3. Close unnecessary applications and browser tabs.\n4. Check Task Manager for processes causing high CPU usage.\n5. Adjust power settings: Control Panel → Power Options → Balanced.\n6. Update BIOS from manufacturer's website.\n7. Consider using a laptop cooling pad.\n8. If laptop is old, internal thermal paste may need replacement (raise hardware ticket).",
     "Medium", "Blocked vents, dust buildup, heavy CPU usage, or dried thermal paste", "15-30 min"],

    ["INC005", "Laptop", "Battery Draining Fast", "Laptop battery lasts only 1-2 hours even with light usage. Used to last 4-5 hours.",
     "1. Check battery health: open CMD → type 'powercfg /batteryreport' → review report.\n2. Reduce screen brightness.\n3. Disable Bluetooth and Wi-Fi when not in use.\n4. Check Power Options and set to 'Battery Saver' mode.\n5. Close background applications in Task Manager.\n6. Disable unnecessary startup programs: Task Manager → Startup tab.\n7. Update Windows and drivers to latest versions.\n8. If battery health is below 40%, request a battery replacement.",
     "Low", "Battery degradation, too many background apps, high brightness, or old battery", "10-20 min"],

    ["INC006", "Laptop", "Keyboard Not Working", "Some or all keys on the laptop keyboard are not responding.",
     "1. Restart the laptop.\n2. Check if Num Lock or Filter Keys are accidentally enabled.\n3. Go to Settings → Accessibility → Keyboard → turn off Filter Keys and Sticky Keys.\n4. Try an external USB keyboard to verify if it's a hardware issue.\n5. Update keyboard driver: Device Manager → Keyboards → Update driver.\n6. Uninstall keyboard driver and restart (Windows will reinstall automatically).\n7. If specific keys don't work, it may be liquid damage — raise hardware ticket.",
     "High", "Driver issue, Filter Keys enabled, liquid damage, or loose ribbon cable", "15-30 min"],

    ["INC007", "Laptop", "Screen Flickering", "Laptop display flickers intermittently, sometimes with horizontal lines.",
     "1. Restart the laptop.\n2. Update display adapter drivers: Device Manager → Display adapters → Update.\n3. Change refresh rate: Settings → Display → Advanced display settings → set to 60Hz.\n4. Disable hardware acceleration in browser settings.\n5. Connect to an external monitor to check if the issue is with the display or GPU.\n6. Boot into Safe Mode — if flickering stops, it's a software/driver issue.\n7. If external monitor works fine, the laptop display panel or cable may be faulty — raise hardware ticket.",
     "Medium", "Outdated display drivers, loose display cable, or failing display panel", "20-40 min"],

    ["INC008", "Laptop", "No Sound", "No audio output from laptop speakers. Volume icon shows no issues.",
     "1. Check volume level and ensure it's not muted (click speaker icon in taskbar).\n2. Right-click speaker icon → Sound settings → check output device is correct.\n3. Try plugging in headphones to test if sound works through headphones.\n4. Run Windows audio troubleshooter: Settings → System → Sound → Troubleshoot.\n5. Update audio driver: Device Manager → Sound → Update driver.\n6. Restart Windows Audio service: services.msc → Windows Audio → Restart.\n7. Uninstall and reinstall audio driver.\n8. If no sound from both speakers and headphones, it may be a hardware issue.",
     "Medium", "Wrong output device selected, muted, driver issue, or hardware failure", "10-25 min"],

    ["INC009", "Laptop", "Webcam Not Working", "Laptop webcam shows black screen or 'No camera found' error in Teams/Zoom.",
     "1. Check if there's a physical privacy shutter/slider covering the webcam.\n2. Check if camera is enabled: Settings → Privacy → Camera → Allow apps to access camera.\n3. Close all other apps that might be using the camera (Zoom, Teams, etc.).\n4. Update camera driver: Device Manager → Cameras → Update driver.\n5. Uninstall camera driver and restart (Windows will reinstall).\n6. Check if antivirus is blocking camera access.\n7. Try the camera in the Windows Camera app to test.\n8. If still not working, try an external USB webcam as a workaround.",
     "Medium", "Privacy shutter closed, permission denied, driver issue, or app conflict", "10-20 min"],

    ["INC010", "Laptop", "Touchpad Not Working", "Laptop touchpad is unresponsive. Cursor does not move.",
     "1. Check if touchpad is disabled — press Fn + F7 (or the touchpad toggle key on your laptop).\n2. Go to Settings → Devices → Touchpad → ensure touchpad is turned on.\n3. If an external mouse is connected, try disconnecting it (some laptops auto-disable touchpad).\n4. Update touchpad driver: Device Manager → Mice and pointing devices → Update.\n5. Uninstall touchpad driver and restart.\n6. Check BIOS settings — ensure touchpad is enabled.\n7. If still not working, use an external mouse and raise a hardware ticket.",
     "Medium", "Touchpad disabled via function key, driver issue, or hardware fault", "10-20 min"],

    ["INC011", "Desktop", "No Display Output", "Desktop is powered on (fans running, lights on) but monitor shows 'No Signal'.",
     "1. Check if the monitor is powered on and set to correct input source (HDMI/DP/VGA).\n2. Ensure the display cable is firmly connected at both ends.\n3. Try a different display cable.\n4. Try connecting to a different monitor.\n5. If the PC has both integrated and dedicated GPU, try the other video port.\n6. Remove and reseat the RAM sticks.\n7. Remove and reseat the GPU.\n8. Try booting with minimal hardware (one RAM stick, no extra drives).\n9. Listen for beep codes — they indicate hardware issues.",
     "High", "Loose cable, wrong input source, GPU failure, or RAM issue", "15-30 min"],

    ["INC012", "Laptop", "Wi-Fi Adapter Missing", "Wi-Fi option is gone from taskbar and network settings. Only Ethernet shows.",
     "1. Check if Airplane Mode is on — turn it off.\n2. Press Fn + Wi-Fi toggle key (varies by laptop model).\n3. Open Device Manager → Network adapters → check if Wi-Fi adapter is listed.\n4. If it shows with a yellow triangle, right-click → Update driver.\n5. If it's missing entirely, try: Action → Scan for hardware changes.\n6. Open CMD (admin) → run: netsh winsock reset\n7. Run: netsh int ip reset\n8. Restart the laptop.\n9. If still missing, the Wi-Fi card may have come loose — raise hardware ticket.",
     "High", "Airplane mode, driver corruption, or loose Wi-Fi card", "15-30 min"],

    ["INC013", "Laptop", "Laptop Lid Close Issue", "Laptop does not go to sleep when lid is closed, or does not wake up when opened.",
     "1. Go to Control Panel → Power Options → Choose what closing the lid does.\n2. Set 'When I close the lid' to 'Sleep' for both battery and plugged in.\n3. If laptop won't wake up: try pressing the power button briefly.\n4. Update chipset and power management drivers.\n5. Disable 'Allow the computer to turn off this device to save power' for USB devices.\n6. Run: powercfg /h on (to ensure hibernation is enabled).\n7. Update BIOS to latest version from manufacturer's website.",
     "Low", "Power settings misconfigured, driver issue, or BIOS setting", "10-15 min"],

    ["INC014", "Laptop", "USB Ports Not Working", "USB devices (mouse, keyboard, flash drive) not recognized when plugged in.",
     "1. Try a different USB port on the laptop.\n2. Try the USB device on another computer to verify the device works.\n3. Restart the laptop.\n4. Device Manager → Universal Serial Bus controllers → Uninstall USB Root Hub → Restart.\n5. Disable USB selective suspend: Power Options → Change plan settings → Change advanced → USB settings.\n6. Run hardware troubleshooter: Settings → Update & Security → Troubleshoot.\n7. Update USB drivers.\n8. If no USB ports work at all, it may be a motherboard issue — raise hardware ticket.",
     "Medium", "Driver issue, power management settings, or hardware fault", "15-25 min"],

    ["INC015", "Laptop", "Docking Station Issues", "Laptop does not detect monitors, USB devices, or network when connected to docking station.",
     "1. Disconnect and reconnect the laptop to the docking station.\n2. Try a different USB-C/Thunderbolt port if available.\n3. Update docking station firmware from manufacturer's website.\n4. Update laptop's Thunderbolt/USB-C drivers.\n5. Check Display settings → Detect monitors manually.\n6. Try connecting monitors directly to the laptop (bypass dock).\n7. Reset the dock: unplug power from dock for 30 seconds.\n8. Try a different power adapter for the dock.\n9. If specific ports on the dock don't work, the dock may be faulty — request replacement.",
     "Medium", "Firmware outdated, driver issue, faulty dock, or cable issue", "20-30 min"],

    # === PASSWORD / ACCOUNT ISSUES (16-30) ===
    ["INC016", "Password", "Forgot Windows Password", "User cannot log into their Windows laptop because they forgot their password.",
     "1. Try any previous passwords you may have used.\n2. If using Azure AD/Microsoft 365, go to https://aka.ms/sspr to reset your password.\n3. Click 'I forgot my password' on the login screen if self-service password reset is enabled.\n4. Use the password reset link sent to your recovery email/phone.\n5. If none of these work, contact IT Helpdesk to reset your Active Directory password.\n6. After reset, log in with the temporary password and change it immediately.\n7. Update the password on your phone's email app and any saved credentials.",
     "High", "Password forgotten after vacation/leave, too frequent changes, or not using password manager", "5-15 min"],

    ["INC017", "Password", "Account Locked Out", "User account is locked out after multiple incorrect password attempts.",
     "1. Wait 30 minutes — most account lockout policies auto-unlock after 30 min.\n2. If urgent, contact IT Helpdesk to manually unlock your account.\n3. Ensure Caps Lock is off when entering the password.\n4. Make sure Num Lock is on if your password contains numbers from the numpad.\n5. Check if someone else might be using your credentials (old sessions on other devices).\n6. After unlocking, change your password immediately if you suspect unauthorized access.\n7. Update the new password on all devices (phone, tablet, etc.).",
     "High", "Caps Lock on, old password cached on other devices, or unauthorized access attempt", "5-10 min"],

    ["INC018", "Password", "Password Expired", "User receives 'Your password has expired and must be changed' message.",
     "1. If at the Windows login screen, click 'OK' and you'll be prompted to set a new password.\n2. Enter your old password, then choose a new password that meets complexity requirements.\n3. Password requirements: minimum 12 characters, mix of uppercase, lowercase, numbers, symbols.\n4. Do not reuse any of your last 10 passwords.\n5. If you cannot change at the login screen, go to https://aka.ms/sspr for self-service reset.\n6. After changing, update the password on your phone, email apps, and VPN client.\n7. Set a reminder to change your password before the next expiry (usually every 90 days).",
     "Medium", "Password expiry policy (90 days), user didn't change before deadline", "5-10 min"],

    ["INC019", "Password", "MFA/2FA Issues", "Unable to complete Multi-Factor Authentication. Not receiving the verification code.",
     "1. Ensure your phone has network/data connectivity.\n2. Check if the authenticator app (Microsoft Authenticator) is installed and set up.\n3. Open the authenticator app and check if the code matches the account.\n4. If using SMS, check if your phone number is correct in your account settings.\n5. Try the 'I can't use my authenticator app right now' option for alternative verification.\n6. If you changed phones, you need to re-register MFA — contact IT Helpdesk.\n7. As a temporary workaround, IT can generate a one-time bypass code.\n8. After resolving, set up backup MFA methods (app + phone + email).",
     "High", "Phone changed, app not synced, network issue, or phone number changed", "10-20 min"],

    ["INC020", "Password", "SSO Not Working", "Single Sign-On (SSO) not working for corporate applications. Getting 'Authentication failed'.",
     "1. Clear browser cache and cookies completely.\n2. Try in an Incognito/Private browsing window.\n3. Ensure you're using the correct corporate email address.\n4. Check if your password was recently changed — SSO tokens may be stale.\n5. Sign out of all Microsoft/Google accounts in the browser and sign in again.\n6. Try a different browser (Edge, Chrome, Firefox).\n7. Check if the issue is with a specific application or all SSO apps.\n8. If VPN is required, ensure you're connected to VPN.\n9. Contact IT if the issue persists — the SSO provider may be having an outage.",
     "Medium", "Stale tokens, cached credentials, browser issue, or SSO provider outage", "10-15 min"],

    ["INC021", "Account", "New Employee Setup", "New employee needs their corporate account and system access set up.",
     "1. Verify the new hire form has been submitted by HR with start date and department.\n2. Create Active Directory account with standard naming convention (first.last).\n3. Assign to appropriate security groups based on department and role.\n4. Create Microsoft 365 mailbox and assign license (E3/E5 as per role).\n5. Set up MFA registration.\n6. Provision access to required applications (Jira, Confluence, Slack, etc.).\n7. Send welcome email with temporary password and setup instructions.\n8. Schedule a 30-min IT orientation session on the first day.\n9. Prepare and ship/deliver laptop with standard image.",
     "Medium", "Standard onboarding process", "30-60 min"],

    ["INC022", "Account", "Employee Offboarding", "Employee is leaving the company. Need to disable access and secure data.",
     "1. Verify the offboarding request from HR with last working date.\n2. On the last day: Disable Active Directory account (do not delete yet).\n3. Reset the password to a random string.\n4. Revoke all MFA tokens and app passwords.\n5. Remove from distribution lists and security groups.\n6. Set up email forwarding to the manager (for 30 days).\n7. Transfer OneDrive/SharePoint files to the manager.\n8. Revoke access to all SaaS applications (Jira, Slack, Salesforce, etc.).\n9. Disable VPN access.\n10. Collect and wipe corporate devices.\n11. Archive mailbox per retention policy.\n12. Delete account after 90 days per policy.",
     "High", "Standard offboarding process — time-sensitive for security", "45-60 min"],

    ["INC023", "Account", "Access Request", "Employee needs access to a specific application or shared drive.",
     "1. Verify the access request is approved by the employee's manager.\n2. Check if the application/resource requires additional approval (data owner, security team).\n3. Identify the correct security group or role for the requested access.\n4. Add the user to the appropriate Active Directory or Azure AD group.\n5. For SaaS apps: provision access through the app's admin portal.\n6. For shared drives: add to the appropriate NTFS/SharePoint permissions group.\n7. Notify the user that access has been granted.\n8. Document the access grant in the ticketing system.\n9. Access will be reviewed during next quarterly access review.",
     "Low", "Standard access provisioning — requires manager approval", "15-30 min"],

    ["INC024", "Password", "Password Policy Help", "User wants to know the password policy and how to create a strong password.",
     "1. Current password policy: Minimum 12 characters.\n2. Must include: uppercase letter (A-Z), lowercase letter (a-z), number (0-9), special character (!@#$%^&*).\n3. Cannot reuse last 10 passwords.\n4. Password expires every 90 days.\n5. Account locks after 5 failed attempts (auto-unlocks after 30 minutes).\n6. Tips for strong passwords: Use a passphrase like 'Coffee$Morning2024!' or 'BlueSky#Rainy42Day'.\n7. Consider using a corporate-approved password manager.\n8. Never share passwords via email, chat, or phone.\n9. Enable MFA for an extra layer of security.",
     "Low", "User education — password awareness", "5 min"],

    ["INC025", "Account", "Shared Mailbox Access", "User needs access to a shared/team mailbox in Outlook.",
     "1. Verify the request is approved by the mailbox owner or manager.\n2. Open Exchange Admin Center or Microsoft 365 Admin.\n3. Navigate to the shared mailbox → Delegation.\n4. Add user as a member with 'Full Access' and/or 'Send As' permission.\n5. The shared mailbox will auto-appear in Outlook within 30-60 minutes.\n6. If it doesn't appear: Outlook → File → Account Settings → Add account.\n7. For Outlook Web: click your profile → Open another mailbox.\n8. Note: Shared mailboxes don't need a separate license if under 50GB.",
     "Low", "Standard access request to shared mailbox", "10-15 min"],

    # === VPN ISSUES (26-35) ===
    ["INC026", "VPN", "Cannot Connect to VPN", "VPN client shows 'Connection failed' or times out when trying to connect.",
     "1. Check your internet connection — try opening a website.\n2. Restart the VPN client application.\n3. Ensure you're using the latest version of the VPN client (GlobalProtect/Cisco AnyConnect).\n4. Check if your credentials are correct — try resetting your password.\n5. Try connecting to a different VPN gateway/server.\n6. Disable any personal firewall or antivirus temporarily.\n7. Restart your laptop and try again.\n8. If on a hotel/airport Wi-Fi, the network may be blocking VPN ports — try using a mobile hotspot.\n9. Flush DNS: Open CMD → ipconfig /flushdns.\n10. If still failing, check the VPN service status page for outages.",
     "High", "Incorrect credentials, network blocking VPN ports, client outdated, or server outage", "15-30 min"],

    ["INC027", "VPN", "VPN Connected But No Access", "VPN shows connected but cannot access internal resources (intranet, file shares, apps).",
     "1. Disconnect and reconnect the VPN.\n2. Check if you can ping internal servers: Open CMD → ping [server-name].\n3. Try accessing internal resources by IP address instead of hostname.\n4. Flush DNS cache: Open CMD (admin) → ipconfig /flushdns.\n5. Release and renew IP: ipconfig /release → ipconfig /renew.\n6. Check if split tunneling is enabled — you may need full tunnel mode.\n7. Verify your VPN profile is correct and you're on the right VPN group.\n8. Check if your account has the necessary network access permissions.\n9. Try accessing from a different network (switch from Wi-Fi to mobile hotspot).",
     "High", "DNS resolution issue, split tunneling, wrong VPN profile, or permission issue", "15-30 min"],

    ["INC028", "VPN", "VPN Keeps Disconnecting", "VPN connection drops every few minutes, requiring frequent reconnection.",
     "1. Check your internet connection stability — run a speed test.\n2. If on Wi-Fi, try moving closer to the router or switch to Ethernet.\n3. Disable Wi-Fi power saving: Device Manager → Network adapter → Properties → Power Management → uncheck 'Allow computer to turn off this device'.\n4. Update VPN client to the latest version.\n5. Update your Wi-Fi adapter driver.\n6. Check if another VPN or security software is conflicting.\n7. Try a different network (mobile hotspot vs. home Wi-Fi).\n8. Adjust VPN idle timeout settings if available.\n9. If the issue persists only on home Wi-Fi, your router may need a firmware update.",
     "Medium", "Unstable internet, Wi-Fi power saving, outdated client, or router issues", "15-25 min"],

    ["INC029", "VPN", "VPN Slow Performance", "Everything works fine but performance is very slow when connected to VPN.",
     "1. Run a speed test with VPN on and off to compare.\n2. Try connecting to a different/closer VPN gateway.\n3. Close bandwidth-heavy applications (video streaming, cloud sync, etc.).\n4. Check if split tunneling is available — enables local internet traffic to bypass VPN.\n5. Reduce video quality in Teams/Zoom calls while on VPN.\n6. If working with large files, transfer them locally before editing.\n7. Check your home internet speed — VPN can only be as fast as your connection.\n8. Try connecting via Ethernet instead of Wi-Fi for more stable speed.\n9. If consistently slow, the VPN gateway may be overloaded — contact IT to check.",
     "Low", "Bandwidth limitations, VPN server load, or all traffic routed through VPN", "10-20 min"],

    ["INC030", "VPN", "VPN Client Installation", "Need to install VPN client on a new laptop or after OS reinstallation.",
     "1. Go to the corporate IT portal / software center.\n2. Download the approved VPN client (GlobalProtect / Cisco AnyConnect / FortiClient).\n3. Run the installer as Administrator.\n4. During setup, enter the VPN server address: vpn.company.com.\n5. Restart the laptop after installation.\n6. Open the VPN client and log in with your corporate credentials.\n7. Complete MFA verification when prompted.\n8. Test by accessing an internal resource (e.g., intranet portal).\n9. If the VPN client is not available in the software center, raise a request to IT.",
     "Low", "New device setup or OS reinstallation", "10-20 min"],

    ["INC031", "VPN", "VPN Certificate Error", "VPN shows 'Certificate error' or 'Invalid certificate' when connecting.",
     "1. Check system date and time — incorrect date/time causes certificate errors.\n2. Sync your clock: Settings → Time & Language → Sync now.\n3. Update the VPN client to the latest version.\n4. Delete the VPN profile and recreate it with the correct server address.\n5. Install the latest corporate root certificates from the IT portal.\n6. If using a personal device, ensure it meets the security compliance requirements.\n7. Clear the VPN client's certificate cache (location varies by client).\n8. As a last resort, uninstall and reinstall the VPN client.\n9. If error mentions a specific certificate, contact IT — the server certificate may need renewal.",
     "Medium", "Incorrect date/time, expired certificate, or missing root CA", "15-25 min"],

    # === NETWORK ISSUES (32-42) ===
    ["INC032", "Network", "No Internet Connection", "Cannot access any websites. Network icon shows 'No Internet Access'.",
     "1. Check if Wi-Fi is turned on (check physical switch or Fn key).\n2. Try a different website to rule out a single-site issue.\n3. Restart your laptop.\n4. Restart your Wi-Fi router (unplug for 30 seconds, plug back in).\n5. Forget the Wi-Fi network and reconnect: Settings → Network → Wi-Fi → Manage known networks.\n6. Run network troubleshooter: Settings → Network → Network troubleshooter.\n7. Reset network: CMD (admin) → netsh winsock reset → netsh int ip reset → restart.\n8. Try connecting to a different Wi-Fi network or use mobile hotspot.\n9. If Ethernet, try a different cable. Check if the port lights blink.",
     "High", "Wi-Fi turned off, router issue, DNS failure, or network adapter problem", "10-30 min"],

    ["INC033", "Network", "Slow Internet Speed", "Internet is working but very slow. Pages take long to load, video calls lag.",
     "1. Run a speed test at speedtest.net — compare with your plan speed.\n2. Disconnect other devices from Wi-Fi that may be using bandwidth.\n3. Move closer to the Wi-Fi router — walls and distance reduce signal.\n4. Restart your router.\n5. Switch from 2.4GHz to 5GHz Wi-Fi band if available (5GHz is faster but shorter range).\n6. Clear browser cache and disable unnecessary browser extensions.\n7. Use Ethernet cable for faster, more stable connection.\n8. Check for Windows updates being downloaded in background.\n9. If speed is consistently low, contact your ISP.",
     "Low", "Wi-Fi interference, too many connected devices, or ISP throttling", "10-20 min"],

    ["INC034", "Network", "Cannot Connect to Wi-Fi", "Wi-Fi networks are visible but cannot connect. Shows 'Can't connect to this network'.",
     "1. Forget the Wi-Fi network and reconnect with the correct password.\n2. Ensure you're entering the correct Wi-Fi password (case-sensitive).\n3. Toggle Wi-Fi off and on.\n4. Toggle Airplane Mode on and off.\n5. Update Wi-Fi driver: Device Manager → Network adapters → Wi-Fi → Update driver.\n6. Reset TCP/IP: CMD (admin) → netsh int ip reset → restart.\n7. Disable and re-enable the Wi-Fi adapter: Device Manager → right-click → Disable/Enable.\n8. Check if MAC address filtering is enabled on the router.\n9. If corporate Wi-Fi, ensure your device certificate is not expired.",
     "Medium", "Wrong password, driver issue, or network configuration mismatch", "15-25 min"],

    ["INC035", "Network", "DNS Resolution Failure", "Getting 'DNS server not responding' or websites not loading but ping by IP works.",
     "1. Flush DNS cache: CMD (admin) → ipconfig /flushdns.\n2. Try changing DNS to Google (8.8.8.8, 8.8.4.4) or Cloudflare (1.1.1.1).\n3. Network settings → Wi-Fi → Properties → IPv4 → Use following DNS servers.\n4. Release/renew IP: ipconfig /release → ipconfig /renew.\n5. Restart the DNS Client service: services.msc → DNS Client → Restart.\n6. Disable IPv6 temporarily: Network adapter Properties → uncheck IPv6.\n7. Restart your router to refresh its DNS cache.\n8. If on VPN, disconnect and try — VPN may be routing DNS incorrectly.",
     "Medium", "ISP DNS issues, corrupted DNS cache, or VPN DNS conflict", "10-20 min"],

    ["INC036", "Network", "Network Printer Not Found", "Cannot find or connect to the network printer. Printer is not listed.",
     "1. Ensure the printer is turned on and connected to the network.\n2. Check the printer's display for its IP address.\n3. Try adding the printer manually: Settings → Printers → Add printer → Add manually → TCP/IP → enter IP.\n4. Ensure you're on the same network as the printer (e.g., not on guest Wi-Fi).\n5. If VPN is connected, the printer may not be reachable — try disconnecting VPN.\n6. Restart the Print Spooler service: services.msc → Print Spooler → Restart.\n7. Install the correct printer driver from the manufacturer's website.\n8. Ask a colleague on the same floor to test the printer.",
     "Medium", "Wrong network, printer offline, missing driver, or VPN interfering", "15-30 min"],

    ["INC037", "Network", "Ethernet Not Working", "Laptop is connected via Ethernet cable but shows 'No network access'.",
     "1. Try a different Ethernet cable.\n2. Try a different Ethernet port on the wall/switch.\n3. Check if the Ethernet adapter is enabled: Device Manager → Network adapters.\n4. Disable and re-enable the Ethernet adapter.\n5. Update the Ethernet driver.\n6. Run: netsh winsock reset and netsh int ip reset in CMD (admin).\n7. Check if the Ethernet port lights are blinking (link and activity LEDs).\n8. If using a USB-C/USB Ethernet adapter, try a different adapter or port.\n9. If the wall port is dead, try a port at a different desk or contact facilities.",
     "Medium", "Bad cable, dead port, disabled adapter, or driver issue", "15-25 min"],

    ["INC038", "Network", "Cannot Access Shared Drive", "Cannot access network shared folders. Getting 'Access denied' or 'Network path not found'.",
     "1. Ensure you're connected to the corporate network (on-site or VPN).\n2. Try accessing the share by IP: \\\\[IP_address]\\sharename instead of hostname.\n3. Check if you have the necessary permissions — ask your manager to verify.\n4. Clear cached credentials: CMD → net use * /delete → try again.\n5. Map the drive again: This PC → Map network drive → enter the path.\n6. Ensure the file server is online — ask colleagues if they can access it.\n7. Restart the Workstation service: services.msc → Workstation → Restart.\n8. If 'Network path not found', the server name may have changed — check with IT.",
     "Medium", "Permission issue, VPN not connected, cached credentials, or server offline", "15-30 min"],

    ["INC039", "Network", "IP Address Conflict", "Getting 'IP address conflict' message. Network connectivity is intermittent.",
     "1. Restart your laptop — it will request a new IP from DHCP.\n2. Release current IP: CMD (admin) → ipconfig /release.\n3. Renew IP: ipconfig /renew.\n4. If using a static IP, check with IT — another device may have the same IP.\n5. Switch to DHCP: Network adapter → Properties → IPv4 → Obtain IP automatically.\n6. Restart your router/switch if you have access.\n7. If the issue persists, the DHCP server may be misconfigured — escalate to Network team.",
     "Medium", "Two devices with same static IP, or DHCP scope exhausted", "10-15 min"],

    ["INC040", "Network", "Teams/Zoom Call Quality", "Video calls on Microsoft Teams or Zoom have poor audio/video quality, freezing, or dropping.",
     "1. Check internet speed — minimum 5 Mbps upload recommended for HD video calls.\n2. Turn off your video if bandwidth is limited — audio alone uses much less bandwidth.\n3. Close other bandwidth-heavy apps during calls.\n4. Use Ethernet instead of Wi-Fi for more stable connection.\n5. If on VPN, try split tunneling or ask IT if Teams can be excluded from VPN.\n6. Lower video quality: Teams → Settings → Reduce video quality.\n7. Close unnecessary browser tabs and applications.\n8. If on Wi-Fi, move closer to the router.\n9. If issue persists on all calls, contact your ISP about your connection quality.",
     "Low", "Low bandwidth, Wi-Fi instability, VPN overhead, or ISP issues", "10-15 min"],

    # === EMAIL ISSUES (41-50) ===
    ["INC041", "Email", "Cannot Send/Receive Emails", "Outlook shows 'Disconnected' or emails are stuck in Outbox.",
     "1. Check your internet connection.\n2. Check Outlook status bar (bottom) — should show 'Connected to: Microsoft Exchange'.\n3. If 'Disconnected' or 'Trying to connect': File → Account Settings → Repair.\n4. Check if your mailbox is full: File → Info → check mailbox size. Clean up if over quota.\n5. For stuck emails in Outbox: Open the email, check for large attachments (>25MB), resize or use OneDrive link.\n6. Try Outlook Web (outlook.office.com) to verify it's not a client-specific issue.\n7. Restart Outlook.\n8. Create a new Outlook profile: Control Panel → Mail → Show Profiles → Add.",
     "High", "Network issue, full mailbox, large attachment, or corrupt Outlook profile", "15-30 min"],

    ["INC042", "Email", "Outlook Keeps Crashing", "Outlook crashes immediately on opening or freezes during use.",
     "1. Start Outlook in Safe Mode: Win+R → outlook.exe /safe.\n2. If it works in Safe Mode, a faulty add-in is the cause — disable add-ins one by one: File → Options → Add-ins → Manage COM Add-ins.\n3. Repair Office: Control Panel → Programs → Microsoft 365 → Modify → Quick Repair.\n4. If Quick Repair doesn't work, try Online Repair (requires internet).\n5. Delete and rebuild Outlook profile: Control Panel → Mail → Show Profiles.\n6. Clear Outlook cache: Close Outlook → delete files in %localappdata%\\Microsoft\\Outlook\\RoamCache.\n7. Ensure Windows and Office are fully updated.\n8. As a workaround, use Outlook Web at outlook.office.com.",
     "Medium", "Faulty add-in, corrupt profile, or outdated Office build", "20-30 min"],

    ["INC043", "Email", "Email Not Syncing on Phone", "Corporate email on mobile phone shows old emails only or won't sync new messages.",
     "1. Check your phone's internet connection.\n2. Open the email app and do a manual refresh (pull down).\n3. Check sync settings: ensure 'Push' or 'Fetch every 15 min' is selected.\n4. Remove and re-add the corporate email account on your phone.\n5. Ensure the Outlook app (or your email app) is updated to the latest version.\n6. If using Microsoft Authenticator for MFA, re-authenticate.\n7. Check if your corporate password was recently changed — update it in the phone's email settings.\n8. Restart your phone.\n9. If using the native Mail app, try the Outlook mobile app instead (recommended).",
     "Medium", "Password changed, sync settings, app outdated, or MFA re-authentication needed", "10-20 min"],

    ["INC044", "Email", "Cannot Open Attachments", "Double-clicking email attachments shows error or asks 'How do you want to open this file?'",
     "1. Check the file type — ensure you have the right software installed (.docx needs Word, .pdf needs PDF reader).\n2. Save the attachment to your desktop first, then open it.\n3. Check if your antivirus is blocking the attachment.\n4. For 'winmail.dat' attachments — the sender is using Outlook Rich Text format. Ask them to resend in HTML format.\n5. If the file is a .zip, right-click → Extract All.\n6. For large files that won't download, check mailbox quota.\n7. Try opening the attachment in Outlook Web (outlook.office.com).\n8. If file type is blocked by policy (.exe, .bat), contact IT for exception or alternative transfer method.",
     "Low", "Missing software, antivirus blocking, corrupt file, or blocked file type", "10-15 min"],

    ["INC045", "Email", "Out of Office Setup", "User wants to set up an automatic Out of Office reply in Outlook.",
     "1. In Outlook Desktop: File → Automatic Replies (Out of Office).\n2. Select 'Send automatic replies' and set the date range.\n3. Write your message for 'Inside My Organization' tab.\n4. Write your message for 'Outside My Organization' tab (if needed).\n5. Include: your return date, who to contact in your absence, and any urgent contact method.\n6. Click OK to activate.\n7. In Outlook Web: Settings (gear icon) → View all Outlook settings → Mail → Automatic replies.\n8. Pro tip: Set a Teams status message too for internal visibility.\n9. The auto-reply will automatically stop on the end date you set.",
     "Low", "Standard request — user education", "5 min"],

    ["INC046", "Email", "Phishing Email Received", "User received a suspicious email that looks like phishing or contains a suspicious link.",
     "1. DO NOT click any links or download attachments in the suspicious email.\n2. DO NOT reply to the email or provide any personal information.\n3. Report the email: In Outlook, click the 'Report Phishing' button (or forward to phishing@company.com).\n4. If you already clicked a link: immediately change your password and enable MFA.\n5. If you entered credentials on a fake site: change your password NOW and contact IT Security.\n6. If you downloaded an attachment: disconnect from network and run a full antivirus scan.\n7. Mark the email as Junk/Spam to help train the filters.\n8. Notify your team if the phishing email appears to target the organization.",
     "High", "Phishing campaign targeting the organization", "10-30 min"],

    ["INC047", "Email", "Distribution List Management", "User wants to create/modify a distribution list or email group.",
     "1. For Microsoft 365 groups: Go to admin.microsoft.com → Groups → Add a group.\n2. Choose group type: Microsoft 365 Group, Distribution list, or Mail-enabled security group.\n3. For distribution list: set name, email address, and add members.\n4. Set who can send to the group (members only vs. anyone).\n5. To modify an existing list: Groups → select group → Members → Add/Remove.\n6. For self-service groups: users can manage in Outlook → People → Groups.\n7. Large changes (bulk add/remove) — provide a CSV file to IT.\n8. All group changes are logged for audit compliance.",
     "Low", "Standard request — requires group owner or IT admin", "15-20 min"],

    ["INC048", "Email", "Email Signature Setup", "User wants to create or update their professional email signature.",
     "1. Open Outlook → File → Options → Mail → Signatures.\n2. Click 'New' to create a new signature.\n3. Recommended format:\n   Full Name | Job Title\n   Department | Company Name\n   Phone: +XX-XXXX-XXXXXX\n   Email: name@company.com\n4. Add the company logo if provided by Marketing/Communications.\n5. Set which signature to use for 'New messages' and 'Replies/forwards'.\n6. Test by sending an email to yourself.\n7. Note: Outlook mobile may need a separate signature setup in the app settings.\n8. Use the corporate signature template if one is available on the intranet.",
     "Low", "Standard request — user education", "5-10 min"],

    ["INC049", "Email", "Calendar Sharing Issues", "Cannot view a colleague's calendar or share own calendar in Outlook.",
     "1. To share your calendar: Calendar → right-click → Sharing Permissions → Add colleague → set permission level.\n2. Permission levels: Availability only, Limited details, Full details, Editor.\n3. To view a shared calendar: Open Calendar → Add Calendar → From Address Book → search colleague.\n4. If you get 'You do not have permission', ask the calendar owner to share with you.\n5. For room/resource calendars: contact IT to set up or modify permissions.\n6. If calendar updates aren't syncing: Ctrl+Shift+F9 to resync in Outlook.\n7. In Teams: you can view free/busy directly when scheduling meetings.\n8. Check if the colleague is in the same organization — external calendar sharing requires admin setup.",
     "Low", "Permission not granted, wrong permission level, or cross-org sharing", "10-15 min"],

    ["INC050", "Email", "Mailbox Full/Over Quota", "Receiving 'Your mailbox is full' or 'Cannot send message, mailbox size exceeded quota' errors.",
     "1. Check mailbox size: Outlook → File → Info → Mailbox Settings.\n2. Empty the Deleted Items folder: right-click → Empty folder.\n3. Empty the Junk Email folder.\n4. Search for emails with large attachments: search 'size:>5MB' → delete or archive.\n5. Archive old emails: File → Info → Archive → select a date range.\n6. Move large attachments to OneDrive and delete the emails.\n7. Empty the Recoverable Items folder: Outlook Web → Deleted Items → Recover items.\n8. Standard mailbox quota is 50GB (E3) or 100GB (E5) — request increase if justified.\n9. Set up auto-archive to manage mailbox size automatically.",
     "Medium", "Large attachments, never cleaned up, or too many retained emails", "15-30 min"],

    # === SOFTWARE ISSUES (51-60) ===
    ["INC051", "Software", "Microsoft Office Not Activating", "Office shows 'Product Activation Required' or features are limited.",
     "1. Ensure you're signed in with your corporate Microsoft account in Office.\n2. File → Account → check if the correct account is listed.\n3. Sign out and sign in again with your corporate email.\n4. If prompted, enter your corporate password and complete MFA.\n5. Check internet connectivity — activation requires internet.\n6. Repair Office: Control Panel → Programs → Microsoft 365 → Modify → Quick Repair.\n7. If Quick Repair doesn't work, try Online Repair.\n8. Run the Microsoft Support and Recovery Assistant (SaRA) tool.\n9. Verify with IT that your account has an Office 365 license assigned.",
     "Medium", "License not assigned, wrong account, or repair needed", "15-30 min"],

    ["INC052", "Software", "Software Installation Request", "User needs a new software application installed on their laptop.",
     "1. Check if the software is available in the Company Portal / Software Center (self-service).\n2. If available, click Install — it will install automatically.\n3. If not available, submit a software request through the IT Service Portal.\n4. Provide: software name, version, business justification, and manager approval.\n5. IT will review for license availability and security compliance.\n6. Free/open-source software also needs IT approval for security review.\n7. Typical turnaround: 1-2 business days for standard software.\n8. For specialized/expensive software: requires budget approval from department head.\n9. After installation, verify the software works and any license activation.",
     "Low", "Standard software provisioning process", "15-30 min"],

    ["INC053", "Software", "Windows Update Issues", "Windows Update is stuck, failing, or showing error codes during update.",
     "1. Restart your laptop and try running Windows Update again.\n2. Run the Windows Update Troubleshooter: Settings → Update & Security → Troubleshoot.\n3. Check available disk space — updates need at least 10GB free.\n4. Clear update cache: Stop Windows Update service → delete contents of C:\\Windows\\SoftwareDistribution → Start service.\n5. Run: DISM /Online /Cleanup-Image /RestoreHealth in CMD (admin).\n6. Run: sfc /scannow in CMD (admin).\n7. Check if VPN is interfering — disconnect VPN and try.\n8. Note the error code and search Microsoft's update troubleshooting page.\n9. If updates consistently fail, a system repair or fresh install may be needed.",
     "Medium", "Low disk space, corrupted update cache, or system file corruption", "20-45 min"],

    ["INC054", "Software", "Application Keeps Crashing", "A specific application crashes frequently or shows 'Not Responding'.",
     "1. Close and restart the application.\n2. Check Task Manager for resource usage — app may need more RAM.\n3. Update the application to the latest version.\n4. Run the application as Administrator: right-click → Run as administrator.\n5. Check Windows Event Viewer for error details: Event Viewer → Windows Logs → Application.\n6. Clear the application's cache/temp files (location varies by app).\n7. Uninstall and reinstall the application.\n8. Check if the issue started after a Windows update — may be a compatibility issue.\n9. As a temporary workaround, try running in Compatibility Mode: right-click .exe → Properties → Compatibility.",
     "Medium", "Insufficient resources, corrupted install, or compatibility issue", "15-30 min"],

    ["INC055", "Software", "Adobe Acrobat Issues", "Cannot open, edit, or print PDF files. Adobe Acrobat shows errors or crashes.",
     "1. Ensure Adobe Acrobat/Reader is updated: Help → Check for Updates.\n2. Set Adobe as default PDF reader: Settings → Default Apps → .pdf → Adobe Acrobat.\n3. Repair the installation: Help → Repair Installation.\n4. If PDF won't open: right-click → Open With → Adobe Acrobat.\n5. For printing issues: try printing to Microsoft Print to PDF first (to test).\n6. Clear Adobe's cache: Preferences → Documents → clear recent files.\n7. Disable Protected Mode: Preferences → Security → uncheck Protected Mode.\n8. If PDFs from a specific source fail, the files may be corrupted — ask the sender to resend.\n9. Uninstall and reinstall Adobe if issues persist.",
     "Low", "Outdated version, corrupt PDF, or conflicting default app", "10-20 min"],

    ["INC056", "Software", "Browser Issues", "Web browser (Chrome/Edge/Firefox) is slow, crashing, or not loading pages correctly.",
     "1. Clear browser cache and cookies: Settings → Privacy → Clear browsing data.\n2. Disable browser extensions one by one to identify the culprit.\n3. Update the browser to the latest version.\n4. Reset browser to default settings: Settings → Reset.\n5. Try a different browser to see if the issue is browser-specific.\n6. Check if proxy settings are correct: Settings → System → Proxy.\n7. Disable hardware acceleration: Settings → System → uncheck 'Use hardware acceleration'.\n8. If specific sites don't load, clear site-specific data or try Incognito mode.\n9. Ensure antivirus isn't interfering with browser traffic.",
     "Low", "Too many extensions, cache bloat, outdated browser, or bad proxy config", "10-20 min"],

    ["INC057", "Software", "Teams Not Working", "Microsoft Teams won't load, crashes, or certain features (chat, calls) not working.",
     "1. Close Teams completely (also from system tray) and restart.\n2. Clear Teams cache: Close Teams → delete contents of %appdata%\\Microsoft\\Teams.\n3. Check if Teams is down: status.office.com or DownDetector.\n4. Update Teams: Settings → Check for updates.\n5. Sign out of Teams and sign back in.\n6. If using web version, clear browser cache.\n7. Check if the issue is with a specific feature (chat works but calls don't → network/VPN issue).\n8. Uninstall and reinstall Teams.\n9. Check your internet speed — Teams needs at least 1.5 Mbps for calls.",
     "Medium", "Cache corruption, service outage, or network issue", "10-20 min"],

    ["INC058", "Software", "Java Runtime Issues", "Application requires Java but showing 'Java not found' or 'Java version error'.",
     "1. Check if Java is installed: CMD → java -version.\n2. If not installed, download from the IT Software Center (not from public sites).\n3. If wrong version, check which version the application requires (Java 8, 11, 17).\n4. Multiple Java versions can coexist — ensure PATH points to the correct one.\n5. Set JAVA_HOME environment variable: System Properties → Environment Variables.\n6. For web applications, Java applets are no longer supported in modern browsers.\n7. After installing, restart the application and the browser.\n8. Some corporate applications may need a specific JRE — check the app documentation.",
     "Low", "Java not installed, wrong version, or PATH not configured", "10-20 min"],

    ["INC059", "Software", "OneDrive Sync Issues", "OneDrive is not syncing files, showing sync errors, or stuck on 'Processing changes'.",
     "1. Check the OneDrive icon in the taskbar for error messages.\n2. Click the OneDrive icon → Help & Settings → View sync problems.\n3. Ensure you have enough cloud storage space (check OneDrive → Settings → Account).\n4. Pause and resume sync: OneDrive icon → Pause syncing → Resume.\n5. Ensure file names don't have invalid characters (#, %, &, etc.).\n6. Check if the file is open — close it and try syncing again.\n7. Files larger than 250GB cannot be synced.\n8. Reset OneDrive: Win+R → %localappdata%\\Microsoft\\OneDrive\\onedrive.exe /reset.\n9. Sign out and sign back in to OneDrive.\n10. Ensure your laptop has enough local disk space for synced files.",
     "Medium", "Invalid file names, file in use, storage full, or sync corruption", "15-30 min"],

    ["INC060", "Software", "Antivirus Alert/Quarantine", "Antivirus software quarantined a file or showing threat alerts.",
     "1. DO NOT ignore antivirus alerts — they may indicate real threats.\n2. Open the antivirus dashboard and review the quarantined file details.\n3. Check the file name and source — was it a download, email attachment, or USB?\n4. If the file is a known false positive (e.g., a corporate tool), request IT to whitelist it.\n5. If the file is suspicious, do not restore it — report to IT Security.\n6. Run a full system scan to check for other threats.\n7. If infected by malware: disconnect from network, run full scan, and contact IT Security immediately.\n8. Change your passwords if the alert mentions a keylogger or credential theft.\n9. Document the alert details for the security team.",
     "High", "Downloaded malware, phishing attachment, or false positive on corporate tool", "15-30 min"],

    # === PRINTER ISSUES (61-67) ===
    ["INC061", "Printer", "Printer Not Printing", "Print job sent but nothing comes out. Printer shows online and ready.",
     "1. Check if the print jobs are stuck in the queue: Settings → Printers → click printer → Open print queue → cancel stuck jobs.\n2. Restart the Print Spooler service: services.msc → Print Spooler → Restart.\n3. Try printing a test page: right-click printer → Printer properties → Print Test Page.\n4. Ensure you selected the correct printer (check default printer).\n5. Check if the printer has paper and toner.\n6. Try printing from a different application.\n7. Remove and re-add the printer: Settings → Printers → Remove → Add.\n8. Update or reinstall the printer driver.\n9. If a network printer, check if you can ping it: CMD → ping [printer-IP].",
     "Medium", "Stuck print queue, wrong printer selected, or spooler service issue", "10-25 min"],

    ["INC062", "Printer", "Paper Jam", "Printer displays 'Paper Jam' error and won't print.",
     "1. Open all printer access doors/trays.\n2. Carefully remove any jammed paper — pull gently in the direction of paper travel.\n3. Check for small pieces of torn paper left inside.\n4. Check all paper trays (main tray, manual feed, duplexer).\n5. Ensure paper in the tray is properly aligned and not overfilled.\n6. Fan the paper stack before loading to prevent sticking.\n7. Use correct paper size and type for the printer.\n8. Close all doors and wait for the printer to reset.\n9. If paper jams are frequent, the rollers may need cleaning — raise a hardware ticket.",
     "Low", "Misaligned paper, overfilled tray, or worn rollers", "5-15 min"],

    ["INC063", "Printer", "Print Quality Issues", "Printouts have streaks, fading, smudging, or missing text/images.",
     "1. Check toner/ink levels and replace if low.\n2. Run the printer's cleaning cycle: Printer Settings → Maintenance → Clean.\n3. Print an alignment/calibration page from printer settings.\n4. For laser printers: remove the toner cartridge, shake gently, reinsert.\n5. Check print settings: ensure correct paper type and quality setting.\n6. For streaks: the drum unit may need replacement.\n7. For fading: check toner density settings.\n8. Try printing from a different application/document to rule out file issues.\n9. If problem persists, the fuser or drum may need replacement — raise a hardware ticket.",
     "Low", "Low toner/ink, dirty print heads, or worn drum unit", "10-20 min"],

    ["INC064", "Printer", "Cannot Install Printer", "Trying to add a printer but Windows cannot find it or driver installation fails.",
     "1. For network printers, get the IP from the printer's display or configuration page.\n2. Add manually: Settings → Printers → Add printer → 'The printer I want isn't listed'.\n3. Select 'Add a printer using TCP/IP' → enter the IP address.\n4. If driver isn't found automatically, download from the manufacturer's website.\n5. For shared printers: \\\\printserver\\printername in File Explorer.\n6. Ensure your user account has permission to install printers (may need admin).\n7. Run 'Print Management' as admin if available.\n8. Check if Group Policy restrictions prevent printer installation.\n9. Ask IT for the printer's full network path and driver package.",
     "Medium", "Wrong IP, missing driver, permission restrictions, or Group Policy", "15-30 min"],

    ["INC065", "Printer", "Printer Offline", "Printer shows as 'Offline' in Windows even though it's powered on.",
     "1. On the printer: check display for errors but confirm it shows 'Ready'.\n2. In Windows: Settings → Printers → click printer → Open print queue → Printer menu → uncheck 'Use Printer Offline'.\n3. Restart the printer (turn off, wait 30 seconds, turn on).\n4. Restart the Print Spooler: services.msc → Print Spooler → Restart.\n5. Check network: ping the printer IP from CMD.\n6. Set the printer as default: right-click → Set as default.\n7. Remove and re-add the printer if still offline.\n8. If the printer's IP changed (DHCP), update the port: Printer Properties → Ports → Configure Port → enter new IP.",
     "Medium", "Printer IP changed, network issue, or Windows marking it offline", "10-20 min"],

    ["INC066", "Printer", "Scan to Email Not Working", "Multifunction printer's scan-to-email feature not working. Scans fail to send.",
     "1. Check if the printer is connected to the network.\n2. Verify the SMTP server settings on the printer: usually smtp.company.com, port 587 or 25.\n3. Check if SMTP authentication credentials are correct.\n4. Ensure the 'From' email address is authorized to send.\n5. Check if the email server requires TLS/SSL — update printer SMTP settings.\n6. Test scanning to a USB drive first — if that works, it's a network/email config issue.\n7. Check firewall rules — ensure the printer can reach the mail server on the required port.\n8. Review the printer's error log for specific failure messages.\n9. Contact IT to verify the SMTP relay allows the printer's IP address.",
     "Medium", "SMTP settings wrong, authentication failure, or firewall blocking", "15-30 min"],

    ["INC067", "Printer", "Duplex Printing Issues", "Cannot print double-sided, or duplex prints are upside down on one side.",
     "1. Check if the printer supports duplex printing.\n2. In print dialog: Printer Properties → check 'Print on Both Sides' or 'Duplex'.\n3. Choose correct flip direction: 'Flip on Long Edge' for portrait, 'Flip on Short Edge' for landscape.\n4. Set duplex as default: Printer Properties → Preferences → Layout → both sides.\n5. If duplex unit is not installed, manual duplex may be available.\n6. For upside-down reverse pages: change binding direction in print properties.\n7. Test with a simple document first to verify orientation.\n8. Update printer driver if duplex option is missing (it may not be a driver feature).",
     "Low", "Wrong flip setting, driver missing duplex option, or duplex unit not installed", "5-15 min"],

    # === MOBILE / REMOTE WORK (68-75) ===
    ["INC068", "Mobile", "Corporate App Not Working on Phone", "Company mobile app (Teams, Outlook, etc.) not working on personal or corporate phone.",
     "1. Ensure the app is updated to the latest version from App Store/Google Play.\n2. Force close the app and reopen.\n3. Clear the app's cache and data: Phone Settings → Apps → [App] → Clear cache.\n4. Check internet connectivity on the phone.\n5. Ensure corporate credentials are correct — re-enter password if recently changed.\n6. Check if MFA is required — complete authentication.\n7. For MDM-managed devices: ensure Intune Company Portal is installed and compliant.\n8. Restart the phone.\n9. Uninstall and reinstall the app if issues persist.\n10. Check if the app requires VPN — connect to corporate VPN first.",
     "Medium", "Outdated app, cached data, changed password, or MDM compliance issue", "10-20 min"],

    ["INC069", "Mobile", "MDM Enrollment Issues", "Unable to enroll personal/corporate device in Mobile Device Management (Intune/MDM).",
     "1. Ensure the device meets minimum OS requirements (iOS 15+ or Android 10+).\n2. Install Microsoft Intune Company Portal from App Store/Google Play.\n3. Open Company Portal → Sign in with corporate account.\n4. Follow the enrollment wizard — grant required permissions.\n5. If enrollment fails: check if your account has an Intune license assigned.\n6. For iOS: ensure a valid MDM profile is installed (Settings → VPN & Device Management).\n7. For Android: ensure Company Portal has all required permissions (storage, phone, etc.).\n8. If previously enrolled, remove old enrollment first: Company Portal → Devices → Remove.\n9. Contact IT if enrollment is blocked by a device compliance policy.",
     "Medium", "Unsupported OS version, missing license, or compliance policy blocking", "15-30 min"],

    ["INC070", "Remote Work", "Remote Desktop Connection Failed", "Cannot connect to office desktop/server via Remote Desktop (RDP). Connection times out.",
     "1. Ensure you're connected to VPN first — RDP to office machines requires VPN.\n2. Check if the remote computer is turned on and not in sleep/hibernate.\n3. Verify the computer name or IP address is correct.\n4. Ensure Remote Desktop is enabled on the target machine: System Properties → Remote → Allow.\n5. Check firewall on the target machine: allow RDP through Windows Firewall.\n6. Try connecting using the IP address instead of hostname.\n7. If 'Authentication error': ensure your account has RDP access permission.\n8. Check if Network Level Authentication (NLA) is causing issues.\n9. Restart the Remote Desktop Service on the target machine if you have access.",
     "High", "VPN not connected, remote PC off, firewall blocking, or permission issue", "15-30 min"],

    ["INC071", "Remote Work", "Home Wi-Fi Setup Help", "Employee working from home needs help optimizing their home network setup.",
     "1. Place the router in a central location, elevated, away from walls and metal objects.\n2. Use 5GHz band for work laptop (faster, less interference).\n3. Change the Wi-Fi channel to reduce interference (usually auto is fine).\n4. Update router firmware from manufacturer's website.\n5. Use Ethernet cable for best stability during video calls.\n6. Minimum recommended speed: 25 Mbps down / 5 Mbps up for remote work.\n7. If coverage is poor, consider a Wi-Fi mesh system or range extender.\n8. Set QoS (Quality of Service) on router to prioritize work laptop.\n9. Separate work and personal device networks if router supports it.\n10. Contact your ISP if speeds are consistently below your plan.",
     "Low", "Router placement, interference, or ISP speed issues", "15-30 min"],

    ["INC072", "Remote Work", "Video Call Background Issues", "Need to set up virtual background or blur background in video calls.",
     "1. Microsoft Teams: Click ••• in the meeting toolbar → Video effects → choose background or upload image.\n2. Zoom: Settings → Background & Effects → choose Virtual Background.\n3. For background blur: select 'Blur' option in the video effects menu.\n4. If virtual background is choppy: your CPU may not support it — use blur instead.\n5. For best results, sit with a plain wall behind you or use good lighting.\n6. Custom backgrounds: use 1920x1080 images for best quality.\n7. If the option is greyed out: update the app or check GPU support requirements.\n8. Corporate-approved backgrounds may be available on the intranet.",
     "Low", "Standard guidance — user education", "5 min"],

    ["INC073", "Mobile", "Two-Factor App Migration", "User got a new phone and needs to transfer Microsoft Authenticator MFA to new device.",
     "1. On the OLD phone (if available): Open Microsoft Authenticator → Settings → Cloud Backup → ensure backup is ON.\n2. On the NEW phone: Install Microsoft Authenticator from App Store/Google Play.\n3. Sign in with the same personal Microsoft account used for backup.\n4. Select 'Begin recovery' and verify your identity.\n5. Accounts should restore from cloud backup.\n6. If old phone is lost/unavailable: contact IT Helpdesk to reset MFA.\n7. IT will clear your MFA methods → you'll re-register at https://aka.ms/mfasetup.\n8. Set up backup verification methods: phone number + another email.\n9. Always keep cloud backup enabled in the Authenticator app.",
     "High", "Phone upgrade/replacement — MFA transfer needed", "15-30 min"],

    ["INC074", "Remote Work", "Workspace Ergonomics", "Employee requesting guidance on ergonomic home office setup.",
     "1. Monitor at arm's length, top of screen at eye level.\n2. Use a separate keyboard and mouse — don't use laptop keyboard for extended work.\n3. Chair height: feet flat on floor, thighs parallel to ground.\n4. Elbows at 90 degrees when typing.\n5. Screen brightness should match room lighting.\n6. Follow 20-20-20 rule: every 20 min, look at something 20 feet away for 20 seconds.\n7. Use a laptop stand or books to elevate the laptop screen.\n8. Equipment available for request: external monitor, keyboard, mouse, headset, laptop stand.\n9. Submit an ergonomic equipment request through the IT Service Portal.",
     "Low", "Standard guidance — employee wellness", "5-10 min"],

    ["INC075", "Remote Work", "Multi-Monitor Setup", "Need help setting up dual/triple monitors with laptop at home or office.",
     "1. Check laptop ports: HDMI, USB-C, DisplayPort, VGA, or Thunderbolt.\n2. For dual monitors, you may need a docking station or USB-C hub.\n3. Connect monitors and go to: Settings → Display.\n4. Click 'Detect' if monitors aren't showing.\n5. Arrange monitors by dragging the display boxes to match physical layout.\n6. Set 'Extend these displays' for additional workspace (not Mirror).\n7. Choose main display: select monitor → check 'Make this my main display'.\n8. Adjust resolution and scaling for each monitor: Settings → Display → Scale and layout.\n9. If using a docking station, update dock firmware and display drivers.\n10. For USB-C monitors, ensure the cable supports video (not all USB-C cables do).",
     "Low", "Cable/adapter compatibility, display settings, or driver/dock issues", "15-25 min"],

    # === SECURITY / COMPLIANCE (76-80) ===
    ["INC076", "Security", "Suspected Data Breach", "User suspects their account has been compromised or data has been leaked.",
     "1. IMMEDIATELY change your password from a trusted device.\n2. Enable or verify MFA is active on your account.\n3. Report to IT Security team at security@company.com or call the security hotline.\n4. DO NOT share the incident details on public channels.\n5. Check your email sent folder and rules for unauthorized forwarding.\n6. Review your Microsoft 365 sign-in history: portal.office.com → My Sign-ins.\n7. Check for any unfamiliar devices in your account: My Account → Devices.\n8. IT Security will investigate: check logs, audit trail, and scope of access.\n9. Do not delete any evidence — IT Security needs it for investigation.\n10. You may be asked to use a different device while yours is checked.",
     "Critical", "Phishing, weak password, or unauthorized access", "Immediate"],

    ["INC077", "Security", "BitLocker Recovery Key", "Laptop showing BitLocker recovery screen after update/hardware change. Needs recovery key.",
     "1. Try restarting the laptop — sometimes a BIOS update triggers BitLocker once.\n2. The recovery key is a 48-digit number.\n3. Check https://aka.ms/myrecoverykey — sign in with your corporate account.\n4. The key is stored in Azure AD / Intune and can be retrieved by IT.\n5. Contact IT Helpdesk with your device name or serial number to retrieve the key.\n6. Enter the recovery key on the BitLocker screen.\n7. After accessing Windows, suspend BitLocker: Control Panel → BitLocker → Suspend.\n8. Restart to verify no more prompts, then resume BitLocker protection.\n9. If this happens repeatedly after BIOS changes, update BIOS to latest and suspend BitLocker before future BIOS updates.",
     "High", "BIOS/firmware update, hardware change, or Secure Boot configuration changed", "10-20 min"],

    ["INC078", "Security", "USB Device Blocked", "Plugging in a USB flash drive shows 'This device is blocked by your organization'.",
     "1. This is a security policy — USB mass storage is restricted on corporate devices.\n2. For file transfer, use approved methods: OneDrive, SharePoint, or corporate file share.\n3. If you need an exception for a specific business need, request through IT Security.\n4. Note: USB keyboards and mice are not blocked — only storage devices.\n5. For presentations on external PCs, upload to OneDrive and access via web.\n6. External hard drives are also blocked under the same policy.\n7. Approved encrypted USB drives may be available — check with IT Security.\n8. Policy exists to prevent data leakage and malware introduction.",
     "Low", "Corporate security policy — data loss prevention", "5-10 min"],

    ["INC079", "Security", "Encryption Setup", "Need to encrypt laptop hard drive or verify encryption status.",
     "1. Check BitLocker status: Control Panel → BitLocker Drive Encryption.\n2. If it shows 'BitLocker on', your drive is already encrypted.\n3. To verify: CMD (admin) → manage-bde -status.\n4. If BitLocker is off: IT must enable it via policy (Intune/SCCM).\n5. Back up your data before enabling encryption (precaution).\n6. Encryption process runs in background — may take 1-2 hours but won't affect usage.\n7. Your recovery key is automatically saved to Azure AD.\n8. For Mac: Check FileVault status → System Preferences → Security → FileVault.\n9. Full disk encryption is mandatory for all corporate devices per security policy.",
     "Medium", "Compliance requirement — mandatory device encryption", "5-15 min"],

    ["INC080", "Security", "Data Classification Help", "User needs guidance on how to classify and label documents/emails properly.",
     "1. Our data classification levels:\n   - PUBLIC: Can be shared externally (marketing materials, press releases).\n   - INTERNAL: For employees only (internal memos, org charts).\n   - CONFIDENTIAL: Restricted to specific teams (financial reports, HR data).\n   - HIGHLY CONFIDENTIAL: Very restricted (M&A data, customer PII, trade secrets).\n2. In Office apps: Home → Sensitivity → choose appropriate label.\n3. In Outlook: Options → Sensitivity → apply before sending.\n4. Labels automatically apply encryption and access controls.\n5. Unsure? Label as CONFIDENTIAL — it's better to over-protect than under-protect.\n6. Customer data (PII, payment info) is always HIGHLY CONFIDENTIAL.\n7. Training materials are available on the intranet: search 'Data Classification Guide'.\n8. Questions about specific data? Contact the Data Protection team.",
     "Low", "Compliance and awareness — data protection", "5-10 min"],
]

# Write data
for row_idx, incident in enumerate(incidents, 2):
    for col_idx, value in enumerate(incident, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

# Adjust column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 60
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 12

# Freeze top row
ws.freeze_panes = 'A2'

# Auto-filter
ws.auto_filter.ref = f"A1:H{len(incidents) + 1}"

# Save
wb.save("src/data/it_helpdesk_kb.xlsx")
print(f"Knowledge base created with {len(incidents)} incidents!")
