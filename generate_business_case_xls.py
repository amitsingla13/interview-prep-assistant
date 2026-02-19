import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Styles
# ============================================================
header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
section_font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
label_font = Font(name='Calibri', size=11)
input_font = Font(name='Calibri', size=11, bold=True, color='1F4E79')
result_font = Font(name='Calibri', size=11, bold=True)
total_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')

header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
input_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
calc_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
section_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
savings_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
total_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
cost_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_row(ws, row, font, fill, cols='BC'):
    for c in cols:
        cell = ws[f'{c}{row}']
        cell.font = font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

def set_cell(ws, row, col, value, font=label_font, fill=None, fmt=None):
    cell = ws[f'{col}{row}']
    cell.value = value
    cell.font = font
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt

# ============================================================
# Sheet 1: Assumptions & Inputs
# ============================================================
ws1 = wb.active
ws1.title = 'Inputs & Assumptions'
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 48
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 14

# Title
ws1.merge_cells('B1:C1')
set_cell(ws1, 1, 'B', 'AI IT Helpdesk — Business Case Calculator', header_font, header_fill)
ws1['C1'].fill = header_fill

# Section: Organization
r = 3
set_cell(ws1, r, 'B', 'ORGANIZATION PARAMETERS', section_font, section_fill)
set_cell(ws1, r, 'C', 'Value', section_font, section_fill)
set_cell(ws1, r, 'D', 'Unit', section_font, section_fill)

inputs = [
    ('Total Employees + Contractors', 40000, ''),
    ('Helpdesk Contacts per Hour', 1000, 'contacts/hr'),
    ('Operating Hours per Day (business hours)', 10, 'hours'),
    ('Average Resolution Time (Human)', 5, 'minutes'),
    ('Working Days per Year', 250, 'days'),
    ('Average L1 Agent Cost (fully loaded, annual)', 45000, '$/year'),
    ('Average Employee Hourly Cost (blended)', 50, '$/hour'),
    ('Productive Hours per Agent per Day', 6, 'hours'),
]

for i, (label, val, unit) in enumerate(inputs):
    row = r + 1 + i
    set_cell(ws1, row, 'B', label, label_font)
    set_cell(ws1, row, 'C', val, input_font, input_fill, '#,##0' if val >= 100 else '0')
    set_cell(ws1, row, 'D', unit, label_font)

# AI Parameters
r2 = r + len(inputs) + 2
set_cell(ws1, r2, 'B', 'AI SOLUTION PARAMETERS', section_font, section_fill)
set_cell(ws1, r2, 'C', 'Value', section_font, section_fill)
set_cell(ws1, r2, 'D', 'Unit', section_font, section_fill)

ai_inputs = [
    ('AI Deflection Rate (L1)', 0.6, '%'),
    ('AI Average Resolution Time', 3, 'minutes'),
    ('AI Off-Hours Deflection Rate', 0.8, '%'),
    ('Off-Hours Shift Premium Multiplier', 1.3, 'x'),
    ('Additional Off-Hours per Day', 14, 'hours'),
]
for i, (label, val, unit) in enumerate(ai_inputs):
    row = r2 + 1 + i
    set_cell(ws1, row, 'B', label, label_font)
    fmt = '0%' if unit == '%' else '0.0' if unit == 'x' else '0'
    set_cell(ws1, row, 'C', val, input_font, input_fill, fmt)
    set_cell(ws1, row, 'D', unit, label_font)

# Solution Cost
r3 = r2 + len(ai_inputs) + 2
set_cell(ws1, r3, 'B', 'SOLUTION COST (ANNUAL)', section_font, section_fill)
set_cell(ws1, r3, 'C', 'Value', section_font, section_fill)
set_cell(ws1, r3, 'D', 'Unit', section_font, section_fill)

cost_inputs = [
    ('OpenAI API Cost (annual)', 200000, '$/year'),
    ('Cloud Hosting (annual)', 40000, '$/year'),
    ('Development & Maintenance FTEs (annual)', 375000, '$/year'),
]
for i, (label, val, unit) in enumerate(cost_inputs):
    row = r3 + 1 + i
    set_cell(ws1, row, 'B', label, label_font)
    set_cell(ws1, row, 'C', val, input_font, input_fill, '#,##0')
    set_cell(ws1, row, 'D', unit, label_font)

total_cost_row = r3 + len(cost_inputs) + 1
set_cell(ws1, total_cost_row, 'B', 'Total Annual Solution Cost', result_font, cost_fill)
set_cell(ws1, total_cost_row, 'C', f'=SUM(C{r3+1}:C{r3+len(cost_inputs)})', result_font, cost_fill, '$#,##0')
set_cell(ws1, total_cost_row, 'D', '$/year', label_font, cost_fill)

# Define named references (row numbers) for the calculations sheet
# Org params: rows 4-11 → C4=employees, C5=contacts/hr, C6=op hrs, C7=res time, C8=working days, C9=agent cost, C10=emp cost, C11=productive hrs
# AI params: rows 14-18 → C14=deflection, C15=AI res time, C16=offhrs deflection, C17=shift premium, C18=offhrs hours
# Cost: rows 21-23 → C21=API, C22=hosting, C23=dev;  C25=total cost

# ============================================================
# Sheet 2: Calculations & Results
# ============================================================
ws2 = wb.create_sheet('Calculations & Results')
ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 28
ws2.column_dimensions['D'].width = 14

ref = "'Inputs & Assumptions'"

# Title
ws2.merge_cells('B1:C1')
set_cell(ws2, 1, 'B', 'AI IT Helpdesk — Calculated Results', header_font, header_fill)
ws2['C1'].fill = header_fill

# ---- Current State ----
r = 3
set_cell(ws2, r, 'B', 'CURRENT STATE (Without AI)', section_font, section_fill)
set_cell(ws2, r, 'C', 'Calculated', section_font, section_fill)
set_cell(ws2, r, 'D', 'Unit', section_font, section_fill)

calcs_current = [
    ('Contacts per Day', f"={ref}!C5*{ref}!C6", '#,##0', 'contacts'),
    ('Total Minutes per Day (employee time)', f"=C4*{ref}!C7", '#,##0', 'minutes'),
    ('Total Hours per Day (employee time)', f"=C5/60", '#,##0.0', 'hours'),
    ('Total Hours per Year (employee time)', f"=C6*{ref}!C8", '#,##0', 'hours'),
    ('Annual Employee Productivity Cost', f"=C7*{ref}!C10", '$#,##0', '$'),
    ('Agent-Hours Needed per Day', f"=C5/60", '#,##0.0', 'hours'),
    ('L1 Agents Needed', f"=ROUNDUP(C9/{ref}!C11,0)", '#,##0', 'agents'),
    ('Annual L1 Staffing Cost', f"=C10*{ref}!C9", '$#,##0', '$'),
]
for i, (label, formula, fmt, unit) in enumerate(calcs_current):
    row = r + 1 + i
    set_cell(ws2, row, 'B', label, label_font, calc_fill)
    set_cell(ws2, row, 'C', formula, result_font, calc_fill, fmt)
    set_cell(ws2, row, 'D', unit, label_font, calc_fill)

# ---- With AI ----
r_ai = r + len(calcs_current) + 2
set_cell(ws2, r_ai, 'B', 'WITH AI HELPDESK', section_font, section_fill)
set_cell(ws2, r_ai, 'C', 'Calculated', section_font, section_fill)
set_cell(ws2, r_ai, 'D', 'Unit', section_font, section_fill)

calcs_ai = [
    ('AI Handles (contacts/day)', f"=C4*{ref}!C14", '#,##0', 'contacts'),
    ('Human Handles (contacts/day)', f"=C4*(1-{ref}!C14)", '#,##0', 'contacts'),
    ('AI Minutes per Day', f"=C{r_ai+1}*{ref}!C15", '#,##0', 'minutes'),
    ('Human Minutes per Day', f"=C{r_ai+2}*{ref}!C7", '#,##0', 'minutes'),
    ('Total Minutes per Day (with AI)', f"=C{r_ai+3}+C{r_ai+4}", '#,##0', 'minutes'),
    ('Total Hours per Year (with AI)', f"=(C{r_ai+5}/60)*{ref}!C8", '#,##0', 'hours'),
    ('Annual Employee Productivity Cost (with AI)', f"=C{r_ai+6}*{ref}!C10", '$#,##0', '$'),
    ('Human Agent-Hours per Day', f"=C{r_ai+4}/60", '#,##0.0', 'hours'),
    ('L1 Agents Needed (with AI)', f"=ROUNDUP(C{r_ai+8}/{ref}!C11,0)", '#,##0', 'agents'),
    ('Annual L1 Staffing Cost (with AI)', f"=C{r_ai+9}*{ref}!C9", '$#,##0', '$'),
]
for i, (label, formula, fmt, unit) in enumerate(calcs_ai):
    row = r_ai + 1 + i
    set_cell(ws2, row, 'B', label, label_font, calc_fill)
    set_cell(ws2, row, 'C', formula, result_font, calc_fill, fmt)
    set_cell(ws2, row, 'D', unit, label_font, calc_fill)

# ---- 24/7 Coverage ----
r_247 = r_ai + len(calcs_ai) + 2
set_cell(ws2, r_247, 'B', '24/7 COVERAGE ANALYSIS', section_font, section_fill)
set_cell(ws2, r_247, 'C', 'Calculated', section_font, section_fill)
set_cell(ws2, r_247, 'D', 'Unit', section_font, section_fill)

calcs_247 = [
    ('Off-Hours Contacts per Day', f"={ref}!C5*{ref}!C18", '#,##0', 'contacts'),
    ('Off-Hours Minutes per Day (human)', f"=C{r_247+1}*{ref}!C7", '#,##0', 'minutes'),
    ('Off-Hours Agent-Hours per Day', f"=C{r_247+2}/60", '#,##0.0', 'hours'),
    ('Off-Hours Agents Needed (human only)', f"=ROUNDUP(C{r_247+3}/{ref}!C11,0)", '#,##0', 'agents'),
    ('Off-Hours Annual Cost (human, with premium)', f"=C{r_247+4}*{ref}!C9*{ref}!C17", '$#,##0', '$'),
    ('Off-Hours AI Handles (%)', f"={ref}!C16", '0%', '%'),
    ('Off-Hours Human Agents (with AI)', f"=ROUNDUP(C{r_247+4}*(1-{ref}!C16),0)", '#,##0', 'agents'),
    ('Off-Hours Annual Cost (with AI, with premium)', f"=C{r_247+7}*{ref}!C9*{ref}!C17", '$#,##0', '$'),
]
for i, (label, formula, fmt, unit) in enumerate(calcs_247):
    row = r_247 + 1 + i
    set_cell(ws2, row, 'B', label, label_font, calc_fill)
    set_cell(ws2, row, 'C', formula, result_font, calc_fill, fmt)
    set_cell(ws2, row, 'D', unit, label_font, calc_fill)

# ---- SAVINGS SUMMARY ----
r_sav = r_247 + len(calcs_247) + 2
set_cell(ws2, r_sav, 'B', 'ANNUAL SAVINGS SUMMARY', section_font, section_fill)
set_cell(ws2, r_sav, 'C', 'Annual $', section_font, section_fill)
set_cell(ws2, r_sav, 'D', '', section_font, section_fill)

savings = [
    ('Employee Productivity Savings', f"=C8-C{r_ai+7}", '$#,##0'),
    ('L1 Staffing Savings', f"=C11-C{r_ai+10}", '$#,##0'),
    ('Agents Reduced', f"=C10-C{r_ai+9}", '#,##0'),
    ('24/7 Coverage Cost Avoidance', f"=C{r_247+5}-C{r_247+8}", '$#,##0'),
]
for i, (label, formula, fmt) in enumerate(savings):
    row = r_sav + 1 + i
    set_cell(ws2, row, 'B', label, label_font, savings_fill)
    set_cell(ws2, row, 'C', formula, result_font, savings_fill, fmt)

# Total
total_row = r_sav + len(savings) + 1
set_cell(ws2, total_row, 'B', 'TOTAL ANNUAL BENEFIT', total_font, total_fill)
set_cell(ws2, total_row, 'C', f"=C{r_sav+1}+C{r_sav+2}+C{r_sav+4}", total_font, total_fill, '$#,##0')

# Solution Cost
set_cell(ws2, total_row + 1, 'B', 'Total Annual Solution Cost', result_font, cost_fill)
set_cell(ws2, total_row + 1, 'C', f"={ref}!C{total_cost_row}", result_font, cost_fill, '$#,##0')

# ROI
set_cell(ws2, total_row + 2, 'B', 'Return on Investment (ROI)', result_font, savings_fill)
set_cell(ws2, total_row + 2, 'C', f"=C{total_row}/C{total_row+1}", result_font, savings_fill, '0.0x')

# Fix the ROI format to show "x" 
ws2[f'C{total_row+2}'].number_format = '0.0"x"'

# Payback in days
set_cell(ws2, total_row + 3, 'B', 'Payback Period', result_font, savings_fill)
set_cell(ws2, total_row + 3, 'C', f"=ROUND(365/C{total_row+2},0)&\" days\"", result_font, savings_fill)

# ============================================================
# Sheet 3: Executive Summary (dashboard-style)
# ============================================================
ws3 = wb.create_sheet('Executive Summary')
ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 22

ws3.merge_cells('B1:E1')
set_cell(ws3, 1, 'B', 'AI IT Helpdesk — Executive Summary', header_font, header_fill)
for c in ['C', 'D', 'E']:
    ws3[f'{c}1'].fill = header_fill

calc_ref = "'Calculations & Results'"

# Side-by-side comparison
set_cell(ws3, 3, 'B', 'Metric', section_font, section_fill)
set_cell(ws3, 3, 'C', 'Current (No AI)', section_font, section_fill)
set_cell(ws3, 3, 'D', 'With AI', section_font, section_fill)
set_cell(ws3, 3, 'E', 'Savings', section_font, section_fill)

exec_rows = [
    ('Contacts per Day', f"={calc_ref}!C4", f"={calc_ref}!C4", '—'),
    ('Avg Resolution Time (min)', f"={ref}!C7", f"={ref}!C15", '—'),
    ('Employee Hours Lost / Year', f"={calc_ref}!C7", f"={calc_ref}!C{r_ai+6}", f"={calc_ref}!C7-{calc_ref}!C{r_ai+6}"),
    ('L1 Agents Required', f"={calc_ref}!C10", f"={calc_ref}!C{r_ai+9}", f"={calc_ref}!C10-{calc_ref}!C{r_ai+9}"),
    ('Annual Productivity Cost', f"={calc_ref}!C8", f"={calc_ref}!C{r_ai+7}", f"={calc_ref}!C8-{calc_ref}!C{r_ai+7}"),
    ('Annual L1 Staffing Cost', f"={calc_ref}!C11", f"={calc_ref}!C{r_ai+10}", f"={calc_ref}!C11-{calc_ref}!C{r_ai+10}"),
    ('24/7 Coverage Cost', f"={calc_ref}!C{r_247+5}", f"={calc_ref}!C{r_247+8}", f"={calc_ref}!C{r_247+5}-{calc_ref}!C{r_247+8}"),
]

fmts = ['#,##0', '0', '#,##0', '#,##0', '$#,##0', '$#,##0', '$#,##0']

for i, (label, cur, ai, sav) in enumerate(exec_rows):
    row = 4 + i
    set_cell(ws3, row, 'B', label, label_font, calc_fill)
    set_cell(ws3, row, 'C', cur, result_font, calc_fill, fmts[i])
    set_cell(ws3, row, 'D', ai, result_font, calc_fill, fmts[i])
    if sav == '—':
        set_cell(ws3, row, 'E', '—', label_font, calc_fill)
    else:
        set_cell(ws3, row, 'E', sav, result_font, savings_fill, fmts[i])

# Totals
tr = 4 + len(exec_rows) + 1
set_cell(ws3, tr, 'B', 'TOTAL ANNUAL BENEFIT', total_font, total_fill)
set_cell(ws3, tr, 'C', '', total_font, total_fill)
set_cell(ws3, tr, 'D', '', total_font, total_fill)
set_cell(ws3, tr, 'E', f"={calc_ref}!C{total_row}", total_font, total_fill, '$#,##0')

set_cell(ws3, tr+1, 'B', 'Total Solution Cost', result_font, cost_fill)
set_cell(ws3, tr+1, 'C', '', label_font, cost_fill)
set_cell(ws3, tr+1, 'D', '', label_font, cost_fill)
set_cell(ws3, tr+1, 'E', f"={ref}!C{total_cost_row}", result_font, cost_fill, '$#,##0')

set_cell(ws3, tr+2, 'B', 'NET ANNUAL BENEFIT', result_font, savings_fill)
set_cell(ws3, tr+2, 'C', '', label_font, savings_fill)
set_cell(ws3, tr+2, 'D', '', label_font, savings_fill)
set_cell(ws3, tr+2, 'E', f"=E{tr}-E{tr+1}", result_font, savings_fill, '$#,##0')

set_cell(ws3, tr+3, 'B', 'ROI', result_font, savings_fill)
set_cell(ws3, tr+3, 'C', '', label_font, savings_fill)
set_cell(ws3, tr+3, 'D', '', label_font, savings_fill)
set_cell(ws3, tr+3, 'E', f"=E{tr}/E{tr+1}", result_font, savings_fill, '0.0"x"')

# Instructions
set_cell(ws3, tr+5, 'B', 'HOW TO USE:', section_font)
set_cell(ws3, tr+6, 'B', '1. Go to "Inputs & Assumptions" sheet', label_font)
set_cell(ws3, tr+7, 'B', '2. Modify any blue-shaded cell to change assumptions', label_font)
set_cell(ws3, tr+8, 'B', '3. All calculations update automatically', label_font)
set_cell(ws3, tr+9, 'B', '4. View results on "Calculations" and this summary sheet', label_font)

# Save
output_path = r'c:\Users\amits\OneDrive\IMP DOCUMENTS\AIDevelopment\codes\Real AI Projects\New Interview preparation Assistant\IT_Helpdesk_Business_Case.xlsx'
wb.save(output_path)
print(f'Saved to {output_path}')
